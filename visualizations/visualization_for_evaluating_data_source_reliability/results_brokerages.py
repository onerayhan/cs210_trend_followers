import os
import re
import openpyxl


def parse_file(file_name):
    pattern = r"(.+)_(\d{2}\.\d{2}\.\d{4})_(\d+)\.txt"
    match = re.match(pattern, file_name)
    if match:
        source = match.group(1).split("\\")[-1]
        date = match.group(2)
        return source, date
    else:
        return None


def find_degerimiz(date):
    workbook = openpyxl.load_workbook("XU100.xlsx")
    worksheet = workbook.active

    for row in worksheet.iter_rows(values_only=True):
        if row[0] == date:
            return round(row[3], 2)

    return None


folder_paths = ["neg", "pos"]
results = []
none_count = 0

for folder_path in folder_paths:
    file_names = os.listdir(folder_path)

    for file_name in file_names:
        file_path = os.path.join(folder_path, file_name)
        result = parse_file(file_path)

        if result:
            source, date = result
            degerimiz = find_degerimiz(date)

            if degerimiz:
                result = {
                    "File Name": file_name,
                    "Folder": folder_path,
                    "Source": source,
                    "Date": date,
                    "Değişim": degerimiz
                }
                results.append(result)
        else:
            none_count += 1

# None değeri sayısını ekrana yazdırma
print("Toplam {} adet None değeri bulundu.".format(none_count))

# Sonuçları dosyaya yazdırma
with open("sonuçlar.txt", "w") as file:
    for result in results:
        file.write(str(result) + "\n")
        file.write("--------------------\n")

print("Sonuçlar başarıyla 'sonuçlar.txt' dosyasına kaydedildi.")

import os
import re
import openpyxl


def parse_file(file_name):
    pattern = r"(.+)_(\d{2}\.\d{2}\.\d{4})_(\d+)\.txt"
    match = re.match(pattern, file_name)
    if match:
        source = match.group(1).split("\\")[-1]
        date = match.group(2)
        return source, date
    else:
        return None


def find_degerimiz(date):
    workbook = openpyxl.load_workbook("XU100.xlsx")
    worksheet = workbook.active

    for row in worksheet.iter_rows(values_only=True):
        if row[0] == date:
            return round(row[3], 2)

    return None


folder_paths = ["neg", "pos"]
results = []
none_count = 0

for folder_path in folder_paths:
    file_names = os.listdir(folder_path)

    for file_name in file_names:
        file_path = os.path.join(folder_path, file_name)
        result = parse_file(file_path)

        if result:
            source, date = result
            degerimiz = find_degerimiz(date)

            if degerimiz:
                result = {
                    "File Name": file_name,
                    "Folder": folder_path,
                    "Source": source,
                    "Date": date,
                    "Değişim": degerimiz
                }
                results.append(result)
        else:
            none_count += 1

# None değeri sayısını ekrana yazdırma
print("Toplam {} adet None değeri bulundu.".format(none_count))

# Sonuçları dosyaya yazdırma
with open("sonuçlarkurumlar.txt", "w") as file:
    for result in results:
        file.write(str(result) + "\n")
        file.write("--------------------\n")

print("Sonuçlar başarıyla 'sonuçlar.txt' dosyasına kaydedildi.")